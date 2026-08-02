# -*- coding: utf-8 -*-
"""二轮定价 · 生态表批量定价 核心逻辑（独立可测试，验证后接入 app.py）

输入：行业表(报名价+券) / 空白生态表(待填P-AE) / 到手价表(红线，可选)
逻辑（用户最终确认）：
  - 百补固定 brand_rate(默认5%)：Q=P*rate, T=P-Q
  - code 统一"倒推红线"：code = ceil0.5(页面价T - 券W - 一轮到手价V)，<0 给 0
  - 20%/30% 是"超code"标记阈值（非上限）；25%/35% 是"超帽"标记阈值
  - 唯一硬约束：二轮到手价 AC <= 一轮到手价 V
  - 无红线新品：按 target_rate 倒推 code = floor0.5(target*(P-W)-Q)
匹配：双键(商品ID+承接SKU) -> 单键唯一价 -> 多价留空标黄+候选 -> 行业无价标"未报名"
列识别：全部按表头列名子串匹配（与列顺序无关）；detect_* / inspect_eco 先识别，
页面可人工修正后再计算（run_eco_pricing 的 cols 参数）；网红名(频道名/渠道名)为可选展示列。
"""
import re
import math
from io import BytesIO

import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment


# ----------------------------- 解析工具 -----------------------------
def parse_price(v):
    """把报名价解析成 float。支持 '594.15 USD' / '$668' / '195.2美金' / 数字。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if pd.notna(v) else None
    s = str(v).strip().replace(",", "")
    if not s or s.lower() in ("nan", "none", "未报名"):
        return None
    m = re.search(r"\d+\.?\d*", s)
    if not m:
        return None
    try:
        return float(m.group())
    except ValueError:
        return None


def parse_coupon(v):
    """把店铺券解析成 float（美金）。优先级见 MEMORY 券解析规则。"""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v) if pd.notna(v) else 0.0
    s = str(v).strip().replace(",", "")
    if not s or s.lower() in ("nan", "none", "无", "-"):
        return 0.0
    # 纯数字
    if re.fullmatch(r"\$?\d+\.?\d*", s):
        return float(re.search(r"\d+\.?\d*", s).group())
    # $X满减 / $X店铺code
    m = re.search(r"\$\s*(\d+\.?\d*)\s*(满减|店铺code|店铺码)", s)
    if m:
        return float(m.group(1))
    # Code-$A-B / 满立减-$A-B / 满$A-B -> 取 B（减免额）
    m = re.search(r"(Code|满立减|满)\s*\$?\s*(\d+\.?\d*)\s*[-减]\s*(\d+\.?\d*)", s, re.IGNORECASE)
    if m:
        return float(m.group(3))
    # X美金 / XUSD
    m = re.search(r"(\d+\.?\d*)\s*(美金|USD|美元)", s, re.IGNORECASE)
    if m:
        return float(m.group(1))
    # 兜底：第一个数字
    m = re.search(r"\d+\.?\d*", s)
    return float(m.group()) if m else 0.0


def ceil05(x):
    return math.ceil(x * 2) / 2.0


def floor05(x):
    return math.floor(x * 2) / 2.0


def norm_id(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = str(v).strip()
    if s.lower() in ("nan", "none"):
        return ""
    if s.endswith(".0"):
        s = s[:-2]
    return s


# ----------------------------- 倾斜钩子（留口子）-----------------------------
def determine_code_rate(商品, 参数, 表现数据=None):
    """过往表现驱动 code 档位倾斜的钩子。当前未启用，返回 None 表示走统一倒推红线逻辑。

    未来：根据 表现数据（如过往GMV/核销率）返回一个建议 code 率或档位偏移。
    """
    return None


# ----------------------------- 列定位 -----------------------------
def _find_col(df, subs, exclude=None, known_idx=None):
    """按子串找列；exclude 用于排除误匹配；找不到回退 known_idx。"""
    for i, c in enumerate(df.columns):
        cs = str(c).replace("\n", " ").strip()
        if exclude and any(e in cs for e in exclude):
            continue
        if any(sub in cs for sub in subs):
            return i
    if known_idx is not None and known_idx < len(df.columns):
        return known_idx
    return None


def detect_hy_columns(df):
    """按表头名自动识别行业表各列。返回 {逻辑列: 列下标(0-based) 或 None}。

    商品ID 与 承接ID 是两列：生态表口径可能用其中任一，故两者都识别、双索引匹配。
    """
    id_c = _find_col(df, ["承接ID", "연계ID"])
    pid_c = _find_col(df, ["商品ID"], exclude=["承接"])
    # 承接SKU：承接ID/商品ID 之后第一个含 SKU 的列（避免误取商品ID前面的 SKU ID 列）
    sku_c = None
    anchor = id_c if id_c is not None else (pid_c if pid_c is not None else -1)
    for i in range(anchor + 1, len(df.columns)):
        if "SKU" in str(df.columns[i]).upper():
            sku_c = i
            break
    if sku_c is None:
        sku_c = _find_col(df, ["承接SKU"])
    price_c = _find_col(df, ["报名价"], exclude=["截图", "询价"])
    coupon_c = _find_col(df, ["券金额"])
    return {"承接ID": id_c, "商品ID": pid_c, "承接SKU": sku_c, "报名价": price_c, "店铺券": coupon_c}


def build_hy_lookups(df_hy, cols=None):
    """从行业表构建：双键价/券、单键唯一价、多价候选。

    cols: detect_hy_columns 的结果（或页面人工确认后的版本）；None 表示自动识别。
    """
    cols = cols or detect_hy_columns(df_hy)
    id_c = cols.get("承接ID")
    pid_c = cols.get("商品ID")
    sku_c = cols.get("承接SKU")
    price_c = cols.get("报名价")
    coupon_c = cols.get("店铺券")

    dual_price, dual_coupon, single_price, multi_price = {}, {}, {}, {}
    coupon_single = {}
    by_id_prices = {}

    for _, r in df_hy.iterrows():
        psku = norm_id(r.iloc[sku_c]) if sku_c is not None else ""
        price = parse_price(r.iloc[price_c]) if price_c is not None else None
        coupon = parse_coupon(r.iloc[coupon_c]) if coupon_c is not None else 0.0
        if price is None:
            continue
        # 同行按 承接ID 和 商品ID 双索引（生态表口径可能用任一，两者都让它命中）
        keys = set()
        if id_c is not None:
            v = norm_id(r.iloc[id_c])
            if v:
                keys.add(v)
        if pid_c is not None:
            v = norm_id(r.iloc[pid_c])
            if v:
                keys.add(v)
        if not keys:
            continue
        for pid in keys:
            dual_price[(pid, psku)] = price
            dual_coupon[(pid, psku)] = coupon
            by_id_prices.setdefault(pid, []).append((price, coupon))

    for pid, lst in by_id_prices.items():
        prices = sorted({round(p, 4) for p, _ in lst})
        if len(prices) == 1:
            single_price[pid] = lst[0][0]
            coupon_single[pid] = lst[0][1]
        else:
            multi_price[pid] = prices

    return {
        "dual_price": dual_price,
        "dual_coupon": dual_coupon,
        "single_price": single_price,
        "coupon_single": coupon_single,
        "multi_price": multi_price,
    }


def detect_id_changes(df_dj, df_hy, dj_cols=None, hy_cols=None):
    """检测承接ID/承接SKU从一轮到二轮是否有变化。

    匹配键：商品ID（跨轮次稳定）。
    一轮来源：到手价表（"商品ID"列实为承接ID口径，承接SKU列）。
    二轮来源：行业表（商品ID、承接ID、承接SKU 三列）。

    返回 (changes_df, stats)：
      changes_df: 每行一个商品，列含 商品ID/商品名/一轮承接ID/二轮承接ID/ID是否变化/
                  一轮承接SKU/二轮承接SKU/SKU是否变化
      stats: {"total", "id_changed", "sku_changed", "both_changed", "unchanged", "new", "r1_only"}
    """
    dj_cols = dj_cols or detect_dj_columns(df_dj)
    hy_cols = hy_cols or detect_hy_columns(df_hy)

    # ── 一轮到手价表索引：商品ID → (承接ID口径值, 承接SKU) ──
    dj_pid_c = dj_cols.get("商品ID")
    dj_sku_c = dj_cols.get("承接SKU")
    r1_map = {}  # 商品ID → {"承接ID": ..., "承接SKU": ...}
    if dj_pid_c is not None:
        for _, r in df_dj.iterrows():
            pid = norm_id(r.iloc[dj_pid_c])
            if not pid:
                continue
            sku = norm_id(r.iloc[dj_sku_c]) if dj_sku_c is not None else ""
            r1_map[pid] = {"承接ID": pid, "承接SKU": sku}  # 到手价表商品ID列=承接ID口径

    # ── 二轮行业表索引：商品ID → (承接ID, 承接SKU, 商品名) ──
    hy_pid_c = hy_cols.get("商品ID")
    hy_id_c = hy_cols.get("承接ID")
    hy_sku_c = hy_cols.get("承接SKU")
    name_c = _find_col(df_hy, ["商品名", "상품명"])
    r2_map = {}  # 商品ID → {"承接ID": ..., "承接SKU": ..., "商品名": ...}
    if hy_pid_c is not None:
        for _, r in df_hy.iterrows():
            pid = norm_id(r.iloc[hy_pid_c])
            if not pid:
                continue
            link_id = norm_id(r.iloc[hy_id_c]) if hy_id_c is not None else pid
            sku = norm_id(r.iloc[hy_sku_c]) if hy_sku_c is not None else ""
            name = str(r.iloc[name_c]).strip() if name_c is not None else ""
            r2_map[pid] = {"承接ID": link_id or pid, "承接SKU": sku, "商品名": name}

    # ── 逐商品对比 ──
    rows = []
    all_pids = sorted(set(list(r1_map.keys()) + list(r2_map.keys())))
    stats = {"total": 0, "id_changed": 0, "sku_changed": 0, "both_changed": 0,
             "unchanged": 0, "new": 0, "r1_only": 0}

    for pid in all_pids:
        r1 = r1_map.get(pid)
        r2 = r2_map.get(pid)
        if r1 and not r2:
            stats["r1_only"] += 1
            continue  # 一轮有、二轮没报名，不列入变化表
        if r2 and not r1:
            stats["new"] += 1
            continue  # 二轮新品，无一轮参照
        # 两轮都有
        stats["total"] += 1
        r1_id = r1["承接ID"]
        r2_id = r2["承接ID"]
        r1_sku = r1["承接SKU"]
        r2_sku = r2["承接SKU"]
        id_changed = (r1_id != r2_id) if (r1_id and r2_id) else False
        sku_changed = (r1_sku != r2_sku) if (r1_sku and r2_sku) else False
        if id_changed:
            stats["id_changed"] += 1
        if sku_changed:
            stats["sku_changed"] += 1
        if id_changed and sku_changed:
            stats["both_changed"] += 1
        if not id_changed and not sku_changed:
            stats["unchanged"] += 1
        # 只记录有变化的行（减少噪音）
        if id_changed or sku_changed:
            rows.append({
                "商品ID": pid,
                "商品名": r2.get("商品名", ""),
                "一轮承接ID": r1_id,
                "二轮承接ID": r2_id,
                "承接ID变化": "⚠️ 已变" if id_changed else "—",
                "一轮承接SKU": r1_sku,
                "二轮承接SKU": r2_sku,
                "承接SKU变化": "⚠️ 已变" if sku_changed else "—",
            })

    changes_df = pd.DataFrame(rows)
    return changes_df, stats


def detect_dj_columns(df):
    """按表头名自动识别到手价表各列。红线价优先「最终价格」，「第一轮到手价」作行级回退。"""
    return {
        "商品ID": _find_col(df, ["商品ID"]),
        "承接SKU": _find_col(df, ["承接SKU"]),
        "红线价": _find_col(df, ["最终价格", "최종할인가"]),
        "一轮到手价": _find_col(df, ["第一轮到手价"]),
        "店铺券": _find_col(df, ["店铺券", "满立减", "스토어 쿠폰"], exclude=["CODE"]),
    }


def build_redline_lookups(df_dj, cols=None):
    """从到手价表构建红线（一轮到手价）：双键 + 单键。红线=最终价格，回退第一轮到手价。

    cols: detect_dj_columns 的结果（或页面人工确认后的版本）；None 表示自动识别。
    """
    if df_dj is None:
        return {"dual": {}, "single": {}, "coupon_dual": {}, "coupon_single": {}}
    cols = cols or detect_dj_columns(df_dj)
    id_c = cols.get("商品ID")
    sku_c = cols.get("承接SKU")
    fp_c = cols.get("红线价")
    r1_c = cols.get("一轮到手价")
    cp_c = cols.get("店铺券")

    dual, single, coupon_dual, coupon_single = {}, {}, {}, {}
    for _, r in df_dj.iterrows():
        pid = norm_id(r.iloc[id_c]) if id_c is not None else ""
        if not pid:
            continue
        psku = norm_id(r.iloc[sku_c]) if sku_c is not None else ""
        rl = parse_price(r.iloc[fp_c]) if fp_c is not None else None
        if rl is None and r1_c is not None:
            rl = parse_price(r.iloc[r1_c])
        cp = parse_coupon(r.iloc[cp_c]) if cp_c is not None else 0.0
        if rl is None:
            continue
        # 双键取最低（最保守红线）
        k = (pid, psku)
        if k not in dual or rl < dual[k]:
            dual[k] = rl
            coupon_dual[k] = cp
        if pid not in single or rl < single[pid]:
            single[pid] = rl
            coupon_single[pid] = cp
    return {"dual": dual, "single": single, "coupon_dual": coupon_dual, "coupon_single": coupon_single}


# ----------------------------- 主流程 -----------------------------
# 生态表列识别：按表头列名匹配（兼容官方完整模板和"只含必需列"的精简模板）
# 输入列：逻辑名 -> (匹配子串, 排除子串)。网红名为可选展示列（官方模板叫"频道名"，行业表叫"渠道名"）
ECO_IN_SPEC = {
    "网红名":   (["频道名", "渠道名", "网红名", "博主", "达人", "채널명"], []),
    "供给类型": (["供给类型"], []),
    "商品ID":   (["商品ID"], []),
    "商品名":   (["商品名"], []),
    "承接SKU":  (["承接SKU ID", "承接SKU"], []),
    "SKU":      (["SKU"], ["承接"]),
    "数量":     (["数量"], []),
}
# 输出列：逻辑键 -> (匹配子串, 排除子串, 缺失时追加的表头名)
ECO_OUT_SPEC = {
    "P":  (["报名原价"], [], "商品报名原价"),
    "Q":  (["百补补贴金额"], [], "百补补贴金额"),
    "R":  (["百补补贴力度"], [], "百补补贴力度"),
    "S":  (["叠加补贴力度"], [], "叠加补贴力度"),
    "T":  (["页面价"], [], "页面价"),
    "V":  (["第一轮到手价"], [], "第一轮到手价"),
    "W":  (["店铺券"], ["CODE"], "店铺券"),
    "Z":  (["code补贴美金"], [], "code补贴美金"),
    "AA": (["code预算"], [], "code预算"),
    "AB": (["折扣率"], [], "折扣率"),
    "AC": (["最终价格"], [], "最终价格"),
    "AD": (["GMV"], [], "GMV"),
    "AE": (["ROI"], [], "ROI"),
}

# 颜色
FILL_RED = PatternFill("solid", fgColor="FFFFE3E3")      # 超价
FILL_ORANGE = PatternFill("solid", fgColor="FFFFE8CC")    # 超code/超帽
FILL_YELLOW = PatternFill("solid", fgColor="FFFFF7CC")    # 待人工/未报名
FILL_GRAY = PatternFill("solid", fgColor="FFF2F2F2")      # 新品无红线


def is_full_supply(s):
    s = str(s)
    return ("全托" in s) or ("海托" in s)


def _match_spec(headers, spec):
    """在 [(列号, 表头名)] 列表上按子串/排除规则匹配逻辑列。返回 {key: 列号 或 None}。"""
    out = {}
    for key, (subs, excl) in spec.items():
        out[key] = None
        for c, name in headers:
            if excl and any(e in name for e in excl):
                continue
            if any(s in name for s in subs):
                out[key] = c
                break
    return out


def read_table(b):
    """解析上传的表格：自动找表头行（前5行里含 商品ID/报名原价/报名价/承接ID 的行）。

    返回 (DataFrame, header_row)。
    """
    df_raw = pd.read_excel(BytesIO(b), header=None)
    header_row = 1
    for i in range(min(5, len(df_raw))):
        rt = " ".join(str(x) for x in df_raw.iloc[i].tolist() if pd.notna(x))
        if "商品ID" in rt or "报名原价" in rt or "报名价" in rt or "承接ID" in rt:
            header_row = i
            break
    df = pd.read_excel(BytesIO(b), header=header_row)
    return df, header_row


def headers_of(df):
    """DataFrame 表头清洗（换行压平、去首尾空格），供页面下拉框展示。"""
    return [str(c).replace("\n", " ").strip() for c in df.columns]


def inspect_eco(eco_bytes):
    """快速读生态表表头行并自动识别各列（read_only 流式读，不加载全表，大表也秒开）。

    返回 (header_row, headers [(1-based列号, 表头名)], detected {逻辑列: 列号 或 None})；
    找不到表头行时返回 (None, [], {})。
    """
    wb = load_workbook(BytesIO(eco_bytes), read_only=True)
    ws = wb.active
    try:
        ws.reset_dimensions()
    except Exception:
        pass
    header_row, headers = None, []
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        rowtext = " ".join(str(x or "") for x in row[:20])
        if "商品ID" in rowtext or "报名原价" in rowtext:
            header_row = ridx
            headers = [(c + 1, str(v).replace("\n", " ").strip())
                       for c, v in enumerate(row) if v is not None and str(v).strip()]
            break
    wb.close()
    if header_row is None:
        return None, [], {}
    return header_row, headers, _match_spec(headers, ECO_IN_SPEC)


def _setup_eco_columns(ws, header_row, in_override=None):
    """按列名识别输入/输出列；输出列缺失时自动追加（精简模板友好）。

    in_override: {逻辑列: 1-based列号}，页面人工确认的列对应，优先于自动识别。
    返回 (in_map, out_map, note_col)。
    """
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None and str(v).strip():
            headers.append((c, str(v).replace("\n", " ").strip()))

    in_map = _match_spec(headers, ECO_IN_SPEC)
    for k, v in (in_override or {}).items():
        if v:
            in_map[k] = v

    out_map = {}
    next_col = ws.max_column + 1
    for key, (subs, excl, header_name) in ECO_OUT_SPEC.items():
        c = None
        for cc, name in headers:
            if excl and any(e in name for e in excl):
                continue
            if any(s in name for s in subs):
                c = cc
                break
        if c is None:
            c = next_col
            ws.cell(header_row, c, header_name)
            next_col += 1
        out_map[key] = c

    # 备注列：已有则复用，否则放到所有已用列之后
    note_col = None
    for cc, name in headers:
        if "匹配备注" in name:
            note_col = cc
            break
    if note_col is None:
        used = [v for v in in_map.values() if v] + list(out_map.values()) + [ws.max_column]
        note_col = max(used) + 1
        ws.cell(header_row, note_col, "匹配备注")
    return in_map, out_map, note_col


def _merged_col_values(ws, col, data_start, max_row):
    """解析某一列的合并单元格：返回 {行号: 值}，合并区内每一行都取首行的值。

    官方生态表的频道名(网红名)按网红合并多行，openpyxl 只在首行返回值，
    其余行为空 —— 用这个补齐，按网红筛选才不会漏掉同组商品。
    """
    merged = {}
    if not col:
        return merged
    for rng in ws.merged_cells.ranges:
        if rng.min_col <= col <= rng.max_col:
            v = ws.cell(rng.min_row, col).value
            for r in range(max(rng.min_row, data_start), min(rng.max_row, max_row) + 1):
                merged[r] = v
    return merged


def run_eco_pricing(df_hy, eco_bytes, df_dj=None, params=None, perf_df=None, cols=None):
    """二轮定价主流程。

    df_hy / df_dj: 用 read_table 解析好的 DataFrame（页面解析一次后传入，避免大文件重复读）。
    eco_bytes: 生态表原始字节（需要 openpyxl 原地填写并导出）。
    cols: {"hy": {...}, "dj": {...}, "eco": {...}} 页面人工确认的列对应；缺省自动识别。
    """
    params = params or {}
    cols = cols or {}
    brand_rate = params.get("brand_rate", 0.05)
    pop_code_flag = params.get("pop_code_flag", 0.20)
    full_code_flag = params.get("full_code_flag", 0.30)
    pop_cap = params.get("pop_cap", 0.25)
    full_cap = params.get("full_cap", 0.35)
    target_rate = params.get("target_rate", 0.24)

    hy = build_hy_lookups(df_hy, cols.get("hy"))
    rl = build_redline_lookups(df_dj, cols.get("dj"))

    # ── 桥接：二轮承接ID → 商品ID（稳定键）──
    # 生态表用二轮承接ID，到手价表用一轮承接ID；承接ID若跨轮变化，直接查会漏。
    # 行业表同行有 商品ID(稳定) + 承接ID(二轮)，建映射让红线兜底能翻译回去。
    hy_cols_resolved = cols.get("hy") or detect_hy_columns(df_hy)
    bridge_id = {}   # 二轮承接ID → 商品ID(稳定)
    bridge_sku = {}  # (二轮承接ID, 二轮承接SKU) → 商品ID(稳定)
    _b_pid = hy_cols_resolved.get("商品ID")
    _b_lid = hy_cols_resolved.get("承接ID")
    _b_sku = hy_cols_resolved.get("承接SKU")
    if _b_pid is not None:
        for _, r in df_hy.iterrows():
            pid = norm_id(r.iloc[_b_pid])
            lid = norm_id(r.iloc[_b_lid]) if _b_lid is not None else ""
            sku = norm_id(r.iloc[_b_sku]) if _b_sku is not None else ""
            if not pid:
                continue
            if lid and lid != pid:
                bridge_id[lid] = pid
            if lid and sku:
                bridge_sku[(lid, sku)] = pid

    wb = load_workbook(BytesIO(eco_bytes))
    ws = wb.active

    # 找表头行（含"商品报名原价"或"商品ID"）
    header_row = None
    for ridx in range(1, min(6, ws.max_row + 1)):
        rowtext = " ".join(str(ws.cell(ridx, c).value or "") for c in range(1, min(20, ws.max_column + 1)))
        if "商品ID" in rowtext or "报名原价" in rowtext:
            header_row = ridx
            break
    if header_row is None:
        header_row = 2
    data_start = header_row + 1

    # 按列名识别输入/输出列（输出列缺失自动追加；页面确认过的列对应优先）
    in_map, out_map, note_col = _setup_eco_columns(ws, header_row, cols.get("eco"))
    if in_map.get("商品ID") is None:
        raise ValueError("生态表里找不到「商品ID」列，请检查表头（可用页面上的模板）")

    # 网红名列若按网红合并了多行，补齐整组取值
    kol_merged = _merged_col_values(ws, in_map.get("网红名"), data_start, ws.max_row)

    preview_rows = []
    stats = {"filled": 0, "multi": 0, "unreg": 0, "over_code": 0, "over_cap": 0,
             "over_price": 0, "new": 0, "total": 0, "dual_hit": 0, "single_hit": 0,
             "bridge_hit": 0}

    for ridx in range(data_start, ws.max_row + 1):
        def _in(key):
            c = in_map.get(key)
            return ws.cell(ridx, c).value if c else None
        kol = kol_merged.get(ridx, _in("网红名"))
        eid = norm_id(_in("商品ID"))
        name = _in("商品名")
        sku_name = _in("SKU")
        esku = norm_id(_in("承接SKU"))
        supply = _in("供给类型")
        qty = parse_price(_in("数量")) or 0.0

        # 跳过完全空行
        if not eid and not name:
            continue
        stats["total"] += 1

        full = is_full_supply(supply)
        code_flag_th = full_code_flag if full else pop_code_flag
        cap_th = full_cap if full else pop_cap

        # ---- 匹配报名价 ----
        note = ""
        status = ""
        P = hy["dual_price"].get((eid, esku))
        W = hy["dual_coupon"].get((eid, esku))
        if P is not None:
            stats["dual_hit"] += 1
        elif eid in hy["single_price"]:
            P = hy["single_price"][eid]
            W = hy["coupon_single"].get(eid, 0.0)
            stats["single_hit"] += 1
        if P is None and eid in hy["multi_price"]:
            cands = hy["multi_price"][eid]
            note = "多价待人工，候选价: " + " / ".join(f"${c:g}" for c in cands)
            status = "🟡 多价待人工"
            stats["multi"] += 1
            _fill_row(ws, ridx, None, out_map, note_col, fill=FILL_YELLOW, note=note)
            preview_rows.append(_preview(kol, eid, name, sku_name, supply, None, None, None, None, None, None, status, note))
            continue
        if P is None:
            status = "🟡 未报名"
            note = "行业表无此商品报名价"
            stats["unreg"] += 1
            _fill_row(ws, ridx, None, out_map, note_col, fill=FILL_YELLOW, note=note)
            preview_rows.append(_preview(kol, eid, name, sku_name, supply, None, None, None, None, None, None, status, note))
            continue

        W = W or 0.0
        # 红线（先直接查，查不到则通过桥接翻译承接ID→商品ID再查）
        V = rl["dual"].get((eid, esku))
        if V is None:
            V = rl["single"].get(eid)
        bridged = False
        if V is None and eid in bridge_id:
            # 承接ID跨轮变化：用稳定商品ID回查一轮到手价
            stable_pid = bridge_id[eid]
            V = rl["single"].get(stable_pid)
            if V is not None:
                bridged = True
                stats["bridge_hit"] += 1
        # 券兜底：行业表无券则用到手价表券
        if (W is None or W == 0.0):
            wfb = rl["coupon_dual"].get((eid, esku)) or rl["coupon_single"].get(eid) or 0.0
            if not wfb and bridged:
                wfb = rl["coupon_single"].get(bridge_id.get(eid, ""), 0.0)
            W = wfb

        # ---- 计算 ----
        Q = P * brand_rate
        T = P - Q
        tilt = determine_code_rate({"id": eid, "sku": esku, "supply": supply},
                                   {"target_rate": target_rate}, perf_df)
        if V is not None:
            need = T - W - V
            code = ceil05(need) if need > 0 else 0.0
            is_new = False
        else:
            code = max(floor05(target_rate * (P - W) - Q), 0.0)
            is_new = True

        if tilt is not None:
            code = tilt  # 预留：倾斜钩子若返回值则覆盖

        S = (Q + code) / (P - W) if (P - W) > 0 else 0.0
        AB = code / T if T > 0 else 0.0
        AC = T - W - code
        AA = qty * code
        AD = qty * T
        AE = AD / AA if AA > 0 else None

        # ---- 标记 ----
        flags = []
        over_code = AB > code_flag_th + 1e-9
        over_cap = S > cap_th + 1e-9
        over_price = (V is not None) and (AC > V + 1e-6)
        if over_code:
            flags.append("超code")
        if over_cap:
            flags.append("超帽")
        if over_price:
            flags.append("超价")

        if is_new:
            status = "⚪ 新品无红线"
            stats["new"] += 1
            fill = FILL_GRAY
        elif over_price:
            status = "🔴 超价放行"
            stats["over_price"] += 1
            fill = FILL_RED
        elif flags:
            status = "⚠️ " + "+".join(flags)
            fill = FILL_ORANGE
            if over_code:
                stats["over_code"] += 1
            if over_cap:
                stats["over_cap"] += 1
        else:
            status = "✅ 压价成功"
            fill = None
            stats["filled"] += 1

        if bridged:
            note = (note + " | " if note else "") + f"承接ID已变(一轮→二轮)，红线经桥接命中(商品ID={bridge_id.get(eid, '')})"

        vals = {"P": P, "Q": Q, "R": brand_rate, "S": S, "T": T,
                "V": V, "W": W, "Z": code, "AA": AA, "AB": AB,
                "AC": AC, "AD": AD, "AE": AE}
        _fill_row(ws, ridx, vals, out_map, note_col, fill=fill, note=note)
        preview_rows.append(_preview(kol, eid, name, sku_name, supply, P, V, T, W, code, AC, status, note, S, AB))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    preview = pd.DataFrame(preview_rows)
    return buf, preview, stats


def _preview(kol, eid, name, sku_name, supply, P, V, T, W, code, AC, status, note="", S=None, AB=None):
    return {
        "网红名": (str(kol).strip()[:14] if kol is not None else ""),
        "商品ID": eid,
        "商品名": (str(name)[:26] if name is not None else ""),
        "SKU": (str(sku_name).strip()[:16] if sku_name is not None else ""),
        "供给类型": (str(supply) if supply is not None else ""),
        "报名价P": round(P, 2) if P is not None else None,
        "红线V": round(V, 2) if V is not None else None,
        "页面价T": round(T, 2) if T is not None else None,
        "券W": round(W, 2) if W is not None else None,
        "code": round(code, 1) if code is not None else None,
        "到手价AC": round(AC, 2) if AC is not None else None,
        "叠加率S": S,
        "code率AB": AB,
        "状态": status,
        "备注": note,
    }


def _fill_row(ws, ridx, vals, out_map, note_col, fill=None, note=""):
    """把计算值写入对应输出列；vals=None 表示待人工行（只上色+备注）。"""
    money_fmt = "0.00"
    rate_fmt = "0.00%"
    code_fmt = "0.0"
    if vals:
        ws.cell(ridx, out_map["P"], round(vals["P"], 2)).number_format = money_fmt
        ws.cell(ridx, out_map["Q"], round(vals["Q"], 2)).number_format = money_fmt
        ws.cell(ridx, out_map["R"], vals["R"]).number_format = rate_fmt
        ws.cell(ridx, out_map["S"], vals["S"]).number_format = rate_fmt
        ws.cell(ridx, out_map["T"], round(vals["T"], 2)).number_format = money_fmt
        if vals["V"] is not None:
            ws.cell(ridx, out_map["V"], round(vals["V"], 2)).number_format = money_fmt
        ws.cell(ridx, out_map["W"], round(vals["W"], 2)).number_format = money_fmt
        ws.cell(ridx, out_map["Z"], vals["Z"]).number_format = code_fmt
        ws.cell(ridx, out_map["AA"], round(vals["AA"], 2)).number_format = money_fmt
        ws.cell(ridx, out_map["AB"], vals["AB"]).number_format = rate_fmt
        ws.cell(ridx, out_map["AC"], round(vals["AC"], 2)).number_format = money_fmt
        ws.cell(ridx, out_map["AD"], round(vals["AD"], 2)).number_format = money_fmt
        if vals["AE"] is not None:
            ws.cell(ridx, out_map["AE"], round(vals["AE"], 2)).number_format = "0.00"
    # 上色（P→AC 区间）
    if fill is not None:
        for c in range(out_map["P"], out_map["AC"] + 1):
            ws.cell(ridx, c).fill = fill
    # 备注
    if note:
        ws.cell(ridx, note_col, note)


# ----------------------------- 输入模板生成 -----------------------------
def _style_template(ws, n_cols):
    """模板表头加粗+灰底，列宽自适应。"""
    fill = PatternFill("solid", fgColor="FFD9D9D9")
    for c in range(1, n_cols + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = 20


def _to_buf(wb):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def make_hy_template():
    """行业表精简模板：含匹配双键（承接ID+商品ID）与定价必需列。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "行业表模板"
    ws.append(["承接ID", "商品ID", "承接SKU ID", "报名价", "店铺券金额"])
    ws.append(["1005012345678901", "1005012345678901", "12000056789012345", 26.59, 2])
    ws.append(["1005012756430808", "1005012137815148", "", 44.91, 0])
    _style_template(ws, 5)
    return _to_buf(wb)


def make_dj_template():
    """到手价表精简模板：红线价来源，必需 3 列（店铺券可选）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "到手价表模板"
    ws.append(["商品ID", "承接SKU ID", "最终价格（红线）", "店铺券"])
    ws.append(["1005012345678901", "12000056789012345", 21.26, 2])
    ws.append(["1005012345678902", "", 34.26, 0])
    _style_template(ws, 4)
    return _to_buf(wb)


def make_eco_template():
    """生态表精简模板：只含必需的信息列，提交后工具自动补齐价格链路列。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "生态表模板"
    ws.append(["网红名（频道名）", "供给类型", "商品ID", "商品名", "承接SKU ID", "SKU", "数量"])
    ws.append(["示例博主A", "POP", "1005012345678901", "示例商品A", "12000056789012345", "黑色", 10])
    ws.append(["示例博主B", "全托管", "1005012345678902", "示例商品B", "", "默认", 5])
    _style_template(ws, 7)
    return _to_buf(wb)
