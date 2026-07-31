"""
网红团购 · 价格链路自动化工具
网红团购一站式平台 - 第一个模块
"""

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import re
import requests
import time
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# 页面配置
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="价格链路自动化 | 网红团购平台",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
# 自定义样式
# ─────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        font-size: 1.6rem;
        font-weight: 700;
        margin-bottom: 0.2rem;
    }
    .sub-header {
        color: #666;
        font-size: 0.9rem;
        margin-bottom: 1.5rem;
    }
    div[data-testid="stSidebar"] {
        background: #fafbfc;
    }
    footer { visibility: hidden; }
    #MainMenu { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 侧边栏：全局参数配置
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ 全局参数")
    st.caption("规则暂定，可随时调整")

    exchange_rate = st.number_input(
        "汇率 (KRW → USD)",
        min_value=1000.0,
        max_value=2000.0,
        value=1550.0,
        step=10.0,
        help="韩元÷此汇率=美金。当前约1550",
    )

    brand_plus_pct = st.slider(
        "百补比例 (Brand+)",
        min_value=0,
        max_value=15,
        value=5,
        step=1,
        format="%d%%",
        help="全托管 / POP半托 brand+ 商品的百补补贴",
    )
    brand_plus_rate = brand_plus_pct / 100.0

    code_cap_pct = st.slider(
        "Code上限 (占原价)",
        min_value=10,
        max_value=35,
        value=20,
        step=1,
        format="%d%%",
        help="网红code补贴占报名原价的上限",
    )
    code_cap_rate = code_cap_pct / 100.0

    total_cap_pct = st.slider(
        "总补贴上限",
        min_value=15,
        max_value=40,
        value=25,
        step=1,
        format="%d%%",
        help="百补+code+店铺券 叠加后不能超过此比例",
    )
    total_cap_rate = total_cap_pct / 100.0

    st.markdown("---")
    st.markdown("**Code自动计算**")

    auto_code = st.checkbox(
        "自动计算Code金额",
        value=True,
        help="根据目标总补贴比例自动倒推code金额（向下取到.5）。关闭则使用表格中已有的code。",
    )

    target_subsidy_pct = st.slider(
        "目标总补贴比例",
        min_value=15,
        max_value=25,
        value=24,
        step=1,
        format="%d%%",
        help="工具按此比例倒推code。设24%留1个点余量，需要时可手动调到25%",
        disabled=not auto_code,
    )
    target_subsidy_rate = target_subsidy_pct / 100.0

    st.markdown("---")
    st.markdown("**校验规则**")

    check_code_round = st.checkbox(
        "Code金额须为整数或.5",
        value=True,
        help="例如 5, 5.5, 10 合法；5.3 不合法",
    )
    check_price_vs_external = st.checkbox(
        "比价：AE最终价 vs 站外最低价",
        value=True,
    )

    st.markdown("---")
    st.markdown("**验价容差**")
    st.caption("报名价 vs 实际页面价的允许偏差")

    tolerance_pass_pct = st.slider(
        "通过阈值 (≤此偏差为正常)",
        min_value=1,
        max_value=10,
        value=5,
        step=1,
        format="%d%%",
        help="偏差在此范围内视为正常价格浮动，不报错",
    )
    tolerance_warn_pct = st.slider(
        "警告阈值 (≤此偏差需留意)",
        min_value=5,
        max_value=30,
        value=15,
        step=1,
        format="%d%%",
        help="偏差超过通过阈值但在警告阈值内，标黄提醒",
    )
    tolerance_abs_usd = st.number_input(
        "绝对差额红线 ($)",
        min_value=10.0,
        max_value=200.0,
        value=50.0,
        step=5.0,
        help="不管百分比多少，差额超过此金额直接标红",
    )

    st.markdown("---")

    enable_naver = st.checkbox(
        "启用 Naver 站外比价",
        value=False,
        help="首次提品需要完整比价时勾选；行业改价后重跑链路时可不勾，节省时间",
    )

    if enable_naver:
        st.markdown("**Naver 比价 API**")
        st.caption("站外最低价查询（密钥只存浏览器会话）")

        naver_id = st.text_input(
            "Client ID",
            value=st.session_state.get("naver_id", ""),
            key="naver_id_input",
            placeholder="Naver 开发者平台 Client ID",
        )
        naver_secret = st.text_input(
            "Client Secret",
            value=st.session_state.get("naver_secret", ""),
            type="password",
            key="naver_secret_input",
            placeholder="Naver 开发者平台 Client Secret",
        )
        if st.button("💾 保存 Naver 密钥", width="stretch"):
            st.session_state["naver_id"] = naver_id.strip()
            st.session_state["naver_secret"] = naver_secret.strip()
            st.success("已保存 ✓")
    else:
        naver_id = ""
        naver_secret = ""

    st.markdown("---")
    st.markdown("**品牌维度（可选）**")
    brand_file = st.file_uploader(
        "上传品牌维度商品信息.xlsx",
        type=["xlsx", "xls"],
        help="上传后可按品牌/类目维度查看GMV分布",
        key="brand_file_uploader",
    )

# ─────────────────────────────────────────────
# 共享工具函数（一轮定价 / 二轮定价 通用）
# ─────────────────────────────────────────────
@st.cache_data
def load_excel(file_bytes):
    """读取Excel，尝试识别表头行；ID类列强制按字符串读避免精度丢失"""
    df_raw = pd.read_excel(BytesIO(file_bytes), header=None)
    # 找到表头行：包含"商品报名原价"或"노미네이션가"的行
    header_row = None
    for i in range(min(5, len(df_raw))):
        row_text = " ".join([str(x) for x in df_raw.iloc[i].tolist() if pd.notna(x)])
        if "报名原价" in row_text or "报名价" in row_text or "노미네이션가" in row_text or "商品ID" in row_text:
            header_row = i
            break

    if header_row is None:
        # 默认第2行(index 1)是表头
        header_row = 1

    # 扫表头识别ID类列（含"ID"或"PID"或"shortkey"），强制按字符串读
    header_cells = df_raw.iloc[header_row].tolist()
    dtype_map = {}
    for col_idx, cell in enumerate(header_cells):
        if pd.isna(cell):
            continue
        cell_str = str(cell).replace("\n", " ").strip().upper()
        if "ID" in cell_str or "PID" in cell_str or "SHORTKEY" in cell_str:
            dtype_map[col_idx] = str

    df = pd.read_excel(BytesIO(file_bytes), header=header_row, dtype=dtype_map)
    return df, header_row


# 列名映射：支持多种可能的列名
COLUMN_ALIASES = {
    "负责人": ["负责人", "메인"],
    "对接人": ["对接人", "서브"],
    "频道名": ["频道名", "채널명"],
    "组": ["组", "조"],
    "团购时间": ["团购时间", "영상 업로드"],
    "是否brand+": ["是否brand+商品", "brand+"],
    "付费商品": ["付费商品", "추천리스트"],
    "供给类型": ["供给类型", "공급 유형"],
    "商品ID": ["商品ID", "상품ID"],
    "商品名": ["商品名", "상품명"],
    "承接SKU_ID": ["承接SKU ID", "진행 상품 SKU"],
    "SKU": ["SKU", "옵션"],
    "数量": ["数量", "진행 수량"],
    "行业反馈库存": ["行业反馈库存"],
    "报名原价": ["商品报名原价", "노미네이션가", "招商口报名价", "报名价"],
    "百补金额": ["百补补贴金额", "빅세이브 할인금액"],
    "百补力度": ["百补补贴力度", "brand+ 할인율"],
    "叠加补贴力度": ["叠加补贴力度", "중복 할인율"],
    "页面价": ["页面价", "상세페이지가격"],
    "店铺券": ["店铺券", "스토어 쿠폰"],
    "店铺券CODE": ["定向发放店铺券CODE", "스토어 쿠폰 네임"],
    "门槛美金": ["门槛美金", "허들"],
    "code金额": ["code补贴美金", "코드금액"],
    "code预算": ["code预算", "코드총예산"],
    "折扣率": ["折扣率", "할인율"],
    "最终价格": ["最终价格", "최종할인가"],
    "GMV": ["GMV"],
    "ROI": ["ROI"],
    "站外美金": ["站外价格（美金）", "국내 최저가(달러)"],
    "站外韩元": ["站外价格（韩元）", "국내 최저가(원)"],
    "站外链接": ["站外比价链接", "국내 최저가 링크"],
    "比价结果": ["比价结果", "가격 비교 결과"],
}


def find_column(df, key):
    """根据别名找到实际列名（按别名优先级搜索，避免短别名误匹配）"""
    aliases = COLUMN_ALIASES.get(key, [key])
    for alias in aliases:
        for col in df.columns:
            col_str = str(col).replace("\n", " ").strip()
            if alias in col_str:
                return col
    return None


def build_formatted_excel(export_df, channel_col=None):
    """按完整价格链路-11.xlsx的精确格式生成Excel（位置编码样式）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    columns = list(export_df.columns)
    n_cols = len(columns)
    n_rows = len(export_df)

    # ── 58列位置编码样式（对标完整价格链路-11.xlsx）──
    # 表头填充色分组（AARRGGBB格式，FF=完全不透明）
    HEADER_COLORS = {
        # A-O (1-15): 黑色
        **{i: "FF000000" for i in range(1, 16)},
        # P-S (16-19): 深灰
        **{i: "FF3F3F3F" for i in range(16, 20)},
        # T-Y (20-25) + AD(30): 橙色
        **{i: "FFC65A14" for i in range(20, 26)},
        30: "FFC65A14",
        # Z-AC (26-29): 红色
        **{i: "FFC10002" for i in range(26, 30)},
        # AE-AH (31-34) + AJ-AO (36-41): 绿色
        **{i: "FF92D04F" for i in range(31, 35)},
        **{i: "FF92D04F" for i in range(36, 42)},
        # AI (35): 无填充（特殊）
        35: None,
        # AP (42): 亮红
        42: "FFFE0300",
        # AQ-AU (43-47): 深蓝
        **{i: "FF00205C" for i in range(43, 48)},
        # AV (48): 浅绿
        48: "FFA9D08D",
        # AW-BF (49-58): 红色
        **{i: "FFC10002" for i in range(49, 59)},
    }

    # 表头不加粗的列（D=4, E=5, F=6, H=8）
    HEADER_NOT_BOLD = {4, 5, 6, 8}

    # 数据行数字格式（按列位置）
    DATA_FORMATS = {
        1: "@",       # A 负责人
        2: "@",       # B 对接人
        3: "@",       # C 频道名
        4: "@",       # D PID
        5: "@",       # E 组
        6: "mm-dd-yy",  # F 团购时间
        10: "@",      # J 商品ID
        11: "@",      # K 商品名
        12: "@",      # L 承接SKU ID
        13: "@",      # M SKU
        14: "General",  # N 数量
        16: "0.00",   # P 报名原价
        17: "$#,##0.00",  # Q 百补金额
        18: "0%",     # R 百补力度
        19: "0.00%",  # S 叠加补贴力度
        20: "$#,##0.00",  # T 页面价
        22: "$#,##0.00",  # V 店铺券
        25: "General",  # Y code金额
        26: "General",  # Z code预算
        27: "0.00%",  # AA 折扣率
        28: "$#,##0.00",  # AB 最终价格
        29: "$#,##0",  # AC GMV
        30: "#,##0.0",  # AD ROI
        31: "$#,##0.00",  # AE 站外美金
        32: "₩#,##0",  # AF 站外韩元
        33: "General",  # AG 站外链接
        34: "General",  # AH 站外截图
        35: "General",  # AI 比价结果
        43: "@",      # AQ shortkey
        48: "@",      # AV 活动ID
    }

    # 数据行水平居中（h=center）的列
    CENTER_H = {16, 19, 20, 25, 26, 27, 28}  # P,S,T,Y,Z,AA,AB

    # 列宽
    COL_WIDTHS = {
        1: 7, 2: 10.6, 3: 8, 4: 13, 5: 8.4, 6: 13,
        10: 16.7, 11: 36.7, 12: 18.3, 13: 25.6,
        16: 11.6, 19: 10.1,
    }
    DEFAULT_WIDTH = 13

    # ── 写表头（第1行）──
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=str(col_name))
        # 字体：白色，大部分加粗，D/E/F/H除外
        is_bold = c_idx not in HEADER_NOT_BOLD
        # AI列(35)特殊：红色字，无填充
        if c_idx == 35:
            cell.font = Font(name="맑은 고딕", size=10, bold=True, color="BE0E1E")
        else:
            cell.font = Font(name="맑은 고딕", size=10, bold=is_bold, color="FFFFFF")
        # 填充色
        fill_color = HEADER_COLORS.get(c_idx, "000000")
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.alignment = header_align
    ws.row_dimensions[1].height = 68

    # ── 写数据（第2行起）──
    data_font = Font(name="맑은 고딕", size=10)
    align_v = Alignment(vertical="center")
    align_hv = Alignment(horizontal="center", vertical="center")

    for r_idx, (_, row) in enumerate(export_df.iterrows(), 2):
        for c_idx, col_name in enumerate(columns, 1):
            val = row[col_name]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = data_font
            cell.alignment = align_hv if c_idx in CENTER_H else align_v
            # 数字格式
            fmt = DATA_FORMATS.get(c_idx)
            if fmt and fmt != "General":
                cell.number_format = fmt
            # 文本格式列确保值为字符串
            if fmt == "@" and val is not None:
                cell.value = str(val)
        ws.row_dimensions[r_idx].height = 24

    # ── 列宽 ──
    for c_idx in range(1, n_cols + 1):
        letter = get_column_letter(c_idx)
        ws.column_dimensions[letter].width = COL_WIDTHS.get(c_idx, DEFAULT_WIDTH)

    # 商品名列(K=11)自动换行
    if n_cols >= 11:
        for r in range(2, n_rows + 2):
            ws.cell(row=r, column=11).alignment = Alignment(
                vertical="center", wrap_text=True
            )

    # ── 合并单元格：A-F 按网红合并 ──
    # 用频道名(C列)ffill确定网红分组边界
    # channel_col 由调用方传入（一轮=col_map 频道名，二轮=map2 频道名）
    if channel_col and channel_col in columns:
        ch_series = export_df[channel_col].ffill()
        groups = []
        start = 0
        for i in range(1, len(ch_series)):
            if ch_series.iloc[i] != ch_series.iloc[start]:
                groups.append((start, i - 1))
                start = i
        groups.append((start, len(ch_series) - 1))

        # 合并 A-F（列1-6）
        merge_col_indices = [i for i in range(1, 7) if i <= n_cols]
        for c_idx in merge_col_indices:
            for g_start, g_end in groups:
                if g_end > g_start:
                    ws.merge_cells(
                        start_row=g_start + 2,
                        start_column=c_idx,
                        end_row=g_end + 2,
                        end_column=c_idx,
                    )
                    ws.cell(row=g_start + 2, column=c_idx).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )

    # ── 冻结窗格：B2（冻结表头行 + A列）──
    ws.freeze_panes = "B2"

    # ── 输出 ──
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─────────────────────────────────────────────
# 主区域
# ─────────────────────────────────────────────
st.markdown('<div class="main-header">💰 价格链路自动化</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">网红团购一站式平台 · 模块一：上传商品表 → 自动计算 → 校验 → 导出</div>', unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 模块切换：一轮定价 / 二轮定价
# ─────────────────────────────────────────────
module_mode = st.radio(
    "定价轮次",
    ["🔵 一轮定价", "🟠 二轮定价 · 价格保护"],
    horizontal=True,
    label_visibility="collapsed",
)

if "二轮" in module_mode:
    st.markdown("### 🟠 二轮定价 · 价格保护")
    st.caption(
        "规则：百补固定 5%，code 帽按供给类型（POP/半托 25%、全托 35%），"
        "唯一约束 = 二轮最终价 ≤ 一轮到手价。只调 code 压价，"
        "给满仍压不下来的标红放行，生成超价清单供与行业谈价。"
    )

    c1, c2 = st.columns(2)
    with c1:
        r1_file = st.file_uploader(
            "① 上传一轮价格链路表（红线价来源）",
            type=["xlsx", "xls"],
            key="r1_file_uploader",
            help="用一轮导出的完整表，系统按商品ID读取每个商品的最低最终价作为红线",
        )
    with c2:
        r2_file = st.file_uploader(
            "② 上传二轮价格链路表（行业新报价）",
            type=["xlsx", "xls"],
            key="r2_file_uploader",
            help="行业二轮提报的表，报名原价通常比一轮高",
        )

    if r1_file is None or r2_file is None:
        st.info("👆 请同时上传一轮表（红线）和二轮表（新报价），系统自动算出每个商品该给多少 code。")
        st.stop()

    # ── 读取两张表 ──
    df1, _ = load_excel(r1_file.getvalue())
    df2, _ = load_excel(r2_file.getvalue())

    def _build_map(dfx):
        m = {}
        for key in COLUMN_ALIASES:
            found = find_column(dfx, key)
            if found:
                m[key] = found
        return m

    map1 = _build_map(df1)
    map2 = _build_map(df2)

    if "商品ID" not in map1:
        st.error("❌ 一轮表缺少「商品ID」列，无法匹配红线价。")
        st.stop()
    if "商品ID" not in map2 or "报名原价" not in map2:
        st.error("❌ 二轮表缺少「商品ID」或「报名原价」列。")
        st.stop()

    # ── 红线价：一轮每个商品的最低最终价（最保守口径）──
    df1["_id"] = df1[map1["商品ID"]].astype(str).str.strip()

    # 优先读一轮表已填的「最终价格」；若是原始空表则按一轮公式还原
    fp1 = pd.to_numeric(df1[map1["最终价格"]], errors="coerce") if map1.get("最终价格") else pd.Series(np.nan, index=df1.index)
    if fp1.notna().sum() == 0:
        if "报名原价" not in map1:
            st.error("❌ 一轮表既没有填「最终价格」，也缺少「报名原价」列，无法还原一轮红线价。")
            st.stop()
        p1 = pd.to_numeric(df1[map1["报名原价"]], errors="coerce")
        b1, s1 = map1.get("是否brand+"), map1.get("供给类型")
        if b1 and s1:
            def _rate1(row):
                sup = str(row.get(s1, "")).strip()
                br = str(row.get(b1, "")).strip().upper()
                if ("POP" in sup or "半托" in sup) and br != "Y":
                    return 0.0
                return brand_plus_rate
            r1 = df1.apply(_rate1, axis=1)
        else:
            r1 = brand_plus_rate
        bb1 = p1 * r1
        pg1 = p1 - bb1
        c1 = map1.get("店铺券")
        cp1 = pd.to_numeric(df1[c1], errors="coerce").fillna(0) if c1 else pd.Series(0.0, index=df1.index)
        cd1_col = map1.get("code金额")
        cd1 = pd.to_numeric(df1[cd1_col], errors="coerce") if cd1_col else None
        if cd1 is not None and cd1.notna().sum() > 0:
            code1 = cd1.fillna(0)
        else:
            code1 = (np.floor((target_subsidy_rate * (p1 - cp1) - bb1) * 2) / 2).clip(lower=0)
        fp1 = pg1 - cp1 - code1
        st.info("ℹ️ 一轮表未填最终价格，已按一轮定价公式自动还原红线价。")

    redline = df1.assign(_fp=fp1).dropna(subset=["_fp"]).groupby("_id")["_fp"].min().to_dict()

    # ── 二轮基础数据 ──
    price2 = map2["报名原价"]
    df2["_id"] = df2[map2["商品ID"]].astype(str).str.strip()
    df2[price2] = pd.to_numeric(df2[price2], errors="coerce")

    coupon2_col = map2.get("店铺券")
    qty2_col = map2.get("数量")
    coupon2 = pd.to_numeric(df2[coupon2_col], errors="coerce").fillna(0) if coupon2_col else pd.Series(0.0, index=df2.index)
    qty2 = pd.to_numeric(df2[qty2_col], errors="coerce").fillna(0) if qty2_col else pd.Series(0.0, index=df2.index)

    # 百补固定 5%（所有品统一，不再区分 brand+）
    df2["_百补力度"] = brand_plus_rate  # 侧栏默认 5%
    df2["_百补金额"] = df2[price2] * df2["_百补力度"]
    df2["_页面价"] = df2[price2] - df2["_百补金额"]
    df2["_一轮红线价"] = df2["_id"].map(redline)

    # 补贴帽按供给类型分档：全托 35%，POP/半托 25%
    supply2 = map2.get("供给类型")
    if supply2:
        df2["_cap_rate"] = df2[supply2].astype(str).apply(
            lambda s: 0.35 if "全托" in s else 0.25
        )
    else:
        df2["_cap_rate"] = 0.25  # 无供给类型列时默认 POP 帽

    # ── 自动算 code：压到红线，但不超补贴上限 ──
    need_code = df2["_页面价"] - coupon2 - df2["_一轮红线价"]
    # code 必须是 0.5 的倍数：压价所需向上取到 0.5（保证不超红线），上限向下取到 0.5
    code_ideal = np.ceil(need_code * 2) / 2
    # 帽顶 = cap_rate × (报名价 − 券) − 百补（cap_rate 按供给类型：全托35%/POP25%）
    max_code_raw = df2["_cap_rate"] * (df2[price2] - coupon2) - df2["_百补金额"]
    code_max = np.floor(max_code_raw * 2) / 2

    # 无一轮红线的新品：按目标补贴比例自动给 code（同一轮逻辑）
    auto_new = (np.floor((target_subsidy_rate * (df2[price2] - coupon2) - df2["_百补金额"]) * 2) / 2).clip(lower=0)

    has_redline = df2["_一轮红线价"].notna()
    code2 = pd.Series(0.0, index=df2.index)
    code2[has_redline] = np.minimum(code_ideal, code_max)[has_redline].clip(lower=0)
    code2[~has_redline] = auto_new[~has_redline]
    df2["_code金额"] = code2

    df2["_最终价格"] = df2["_页面价"] - coupon2 - df2["_code金额"]
    df2["_超价金额"] = df2["_最终价格"] - df2["_一轮红线价"]

    def _status(row):
        if pd.isna(row["_一轮红线价"]):
            return "⚪ 新品·无一轮红线"
        if row["_最终价格"] <= row["_一轮红线价"] + 1e-9:
            return "✅ 压价成功"
        return "🔴 超价放行"
    df2["_状态"] = df2.apply(_status, axis=1)

    df2["_code预算"] = qty2 * df2["_code金额"]
    df2["_GMV"] = qty2 * df2["_页面价"]
    df2["_ROI"] = np.where(df2["_code预算"] > 0, df2["_GMV"] / df2["_code预算"], np.nan)
    denom2 = df2[price2] - coupon2
    df2["_叠加补贴"] = np.where(denom2 > 0, (df2["_百补金额"] + df2["_code金额"]) / denom2, 0)
    df2["_折扣率"] = np.where(df2["_页面价"] > 0, df2["_code金额"] / df2["_页面价"], 0)

    # ── 概览指标 ──
    st.markdown("---")
    n_ok = int((df2["_状态"] == "✅ 压价成功").sum())
    n_over = int((df2["_状态"] == "🔴 超价放行").sum())
    n_new = int((df2["_状态"] == "⚪ 新品·无一轮红线").sum())
    g1, g2, g3, g4, g5 = st.columns(5)
    g1.metric("二轮商品数", len(df2))
    g2.metric("✅ 压价成功", n_ok)
    g3.metric("🔴 超价放行", n_over)
    g4.metric("⚪ 新品无红线", n_new)
    g5.metric("Code总预算", f"${df2['_code预算'].sum():,.0f}")

    if n_over > 0:
        st.warning(f"有 {n_over} 个商品即使给满 code 也压不到一轮红线价，已标红放行，请拿下方超价清单与行业谈价。")

    # ── 结果表（超价行标红）──
    st.markdown("### 📋 二轮定价结果")
    name2 = map2.get("商品名")
    supply2_col = map2.get("供给类型")
    df_show = pd.DataFrame({
        "商品ID": df2["_id"],
        "商品名": (df2[name2].astype(str).str.slice(0, 28) if name2 else pd.Series("", index=df2.index)),
        "供给类型": (df2[supply2_col].astype(str) if supply2_col else pd.Series("", index=df2.index)),
        "帽": (df2["_cap_rate"] * 100).astype(int).astype(str) + "%",
        "一轮红线价": df2["_一轮红线价"].round(2),
        "二轮报名原价": df2[price2].round(2),
        "二轮页面价": df2["_页面价"].round(2),
        "店铺券": coupon2.round(2),
        "code金额": df2["_code金额"],
        "二轮最终价": df2["_最终价格"].round(2),
        "超价金额": df2["_超价金额"].apply(lambda x: f"+{x:.2f}" if pd.notna(x) and x > 0.001 else ""),
        "状态": df2["_状态"],
    })

    def _hl(row):
        if row["状态"] == "🔴 超价放行":
            return ["background-color:#ffe3e3"] * len(row)
        if row["状态"] == "⚪ 新品·无一轮红线":
            return ["background-color:#f2f2f2"] * len(row)
        return [""] * len(row)

    _money = st.column_config.NumberColumn(format="%.2f")
    st.dataframe(
        df_show.style.apply(_hl, axis=1),
        width="stretch",
        hide_index=True,
        height=420,
        column_config={
            "一轮红线价": _money,
            "二轮报名原价": _money,
            "二轮页面价": _money,
            "店铺券": _money,
            "code金额": st.column_config.NumberColumn(format="%.1f"),
            "二轮最终价": _money,
        },
    )

    # ── 超价清单（给行业谈价用）──
    if n_over > 0:
        st.markdown("### 🔴 超价清单（与行业谈价用）")
        st.caption("这些商品已给满补贴上限允许的 code，仍高于一轮红线，需行业下调报名原价或接受超价。")
        ov = df2[df2["_状态"] == "🔴 超价放行"]
        df_over = pd.DataFrame({
            "商品ID": ov["_id"],
            "商品名": (ov[name2].astype(str).str.slice(0, 30) if name2 else pd.Series("", index=ov.index)),
            "一轮红线价": ov["_一轮红线价"].round(2),
            "二轮页面价": ov["_页面价"].round(2),
            "已给code(上限)": ov["_code金额"],
            "二轮最终价": ov["_最终价格"].round(2),
            "超价金额": ov["_超价金额"].round(2),
        })
        st.dataframe(
            df_over,
            width="stretch",
            hide_index=True,
            column_config={
                "一轮红线价": _money,
                "二轮页面价": _money,
                "已给code(上限)": st.column_config.NumberColumn(format="%.1f"),
                "二轮最终价": _money,
                "超价金额": _money,
            },
        )

    # ── 导出 ──
    st.markdown("---")
    st.markdown("### 📥 导出")
    e1, e2 = st.columns(2)
    with e1:
        export2 = df2.copy()
        if map2.get("百补金额"):
            export2[map2["百补金额"]] = export2["_百补金额"]
        if map2.get("百补力度"):
            _rv = export2["_百补力度"]
            export2.loc[_rv > 0, map2["百补力度"]] = _rv[_rv > 0]
        if map2.get("页面价"):
            export2[map2["页面价"]] = export2["_页面价"]
        if map2.get("code金额"):
            export2[map2["code金额"]] = export2["_code金额"]
        if map2.get("最终价格"):
            export2[map2["最终价格"]] = export2["_最终价格"]
        if map2.get("code预算"):
            export2[map2["code预算"]] = export2["_code预算"]
        if map2.get("GMV"):
            export2[map2["GMV"]] = export2["_GMV"]
        if map2.get("ROI"):
            export2[map2["ROI"]] = export2["_ROI"]
        if map2.get("叠加补贴力度"):
            export2[map2["叠加补贴力度"]] = export2["_叠加补贴"]
        if map2.get("折扣率"):
            export2[map2["折扣率"]] = export2["_折扣率"]
        export2 = export2.drop(columns=[c for c in export2.columns if c.startswith("_")])
        buf2 = build_formatted_excel(export2, map2.get("频道名"))
        st.download_button(
            label="📥 下载二轮定价表（已自动压价）",
            data=buf2,
            file_name="二轮定价_压价完成.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
    with e2:
        if n_over > 0:
            buf3 = BytesIO()
            df_over.to_excel(buf3, index=False, engine="openpyxl")
            buf3.seek(0)
            st.download_button(
                label="📥 下载超价清单",
                data=buf3,
                file_name="二轮定价_超价清单.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )

    st.markdown("---")
    st.caption("网红团购一站式平台 · 价格链路模块（二轮定价）| 数据仅在本地浏览器处理，不上传任何服务器")
    st.stop()



# ─────────────────────────────────────────────
# 模块页签：上传 / 计算 / 校验 / 分析 / 导出
# ─────────────────────────────────────────────
tab_upload, tab_calc, tab_valid, tab_analyze, tab_export = st.tabs(
    ["📤 上传", "🧮 计算", "✅ 校验", "📊 分析", "📥 导出"]
)

with tab_upload:
    uploaded_file = st.file_uploader(
        "上传价格链路表 (.xlsx)",
        type=["xlsx", "xls"],
        help="支持你现有的价格链路表格式，系统会自动识别列",
    )
    if uploaded_file is None:
        st.info("👆 上传你的价格链路 Excel 表，系统会自动完成所有计算和校验。")
        st.markdown("""
        **工具会自动完成：**
        - 百补金额、页面价、最终价格、GMV、ROI 等全部公式计算
        - Code金额校验（整数/.5）
        - 叠加补贴是否超上限
        - 站外比价（韩元→美金换算 + 高低判断）
        - 异常标红提醒
        - 一键导出完整表格
        """)

if uploaded_file is None:
    for _t in (tab_calc, tab_valid, tab_analyze, tab_export):
        with _t:
            st.info("请先在「上传」页签上传价格链路表，上传后自动开始计算。")
    st.stop()

# ─────────────────────────────────────────────
# Step 2: 读取并识别列
# ─────────────────────────────────────────────


df, header_row = load_excel(uploaded_file.getvalue())





# 建立列映射
col_map = {}
for key in COLUMN_ALIASES:
    found = find_column(df, key)
    if found:
        col_map[key] = found
# ─────────────────────────────────────────────
# Step 3: 自动计算
# ─────────────────────────────────────────────
# 确保数值列为数字类型
numeric_cols = ["报名原价", "数量", "店铺券", "code金额", "站外韩元", "站外美金"]
for key in numeric_cols:
    if key in col_map:
        col = col_map[key]
        df[col] = pd.to_numeric(df[col], errors="coerce")

# 获取关键列名
col_price = col_map.get("报名原价")
col_qty = col_map.get("数量")
col_coupon = col_map.get("店铺券")
col_code = col_map.get("code金额")
col_krw = col_map.get("站外韩元")
col_usd_ext = col_map.get("站外美金")
col_brand = col_map.get("是否brand+")
col_supply = col_map.get("供给类型")
col_name = col_map.get("商品名")
col_result = col_map.get("比价结果")

if not col_price:
    st.error("❌ 未找到「报名原价」列，请检查表格格式。")
    st.stop()

# 计算各列
# 百补力度：根据供给类型和brand+判断
if col_brand and col_supply:
    def get_brand_rate(row):
        supply = str(row.get(col_supply, "")).strip()
        brand = str(row.get(col_brand, "")).strip().upper()
        # POP/半托管 且 非brand+ → 无百补
        if ("POP" in supply or "半托" in supply) and brand != "Y":
            return 0.0
        return brand_plus_rate
    df["_百补力度"] = df.apply(get_brand_rate, axis=1)
else:
    df["_百补力度"] = brand_plus_rate

# 百补金额 = 原价 × 百补力度
df["_百补金额"] = df[col_price] * df["_百补力度"]

# 页面价 = 原价 - 百补金额
df["_页面价"] = df[col_price] - df["_百补金额"]

# 店铺券（默认为0）
if col_coupon:
    coupon_values = df[col_coupon].fillna(0)
else:
    coupon_values = pd.Series(0, index=df.index)

# Code金额：自动计算 或 使用表格已有值
if auto_code:
    # 倒推公式: code = 目标比例 × (原价 - 店铺券) - 百补金额
    raw_code = target_subsidy_rate * (df[col_price] - coupon_values) - df["_百补金额"]
    # 向下取到最近的0.5（例如 5.75 → 5.5, 2.3 → 2.0, 4.99 → 4.5）
    code_values = (np.floor(raw_code * 2) / 2).clip(lower=0)
else:
    if col_code:
        code_values = df[col_code].fillna(0)
    else:
        code_values = pd.Series(0, index=df.index)

# 最终价格 = 页面价 - 店铺券 - code
df["_最终价格"] = df["_页面价"] - coupon_values - code_values

# Code预算 = 数量 × code金额
if col_qty:
    df["_code预算"] = df[col_qty].fillna(0) * code_values
else:
    df["_code预算"] = code_values

# GMV = 数量 × 页面价
if col_qty:
    df["_GMV"] = df[col_qty].fillna(0) * df["_页面价"]
else:
    df["_GMV"] = df["_页面价"]

# ROI = GMV / code预算
df["_ROI"] = np.where(df["_code预算"] > 0, df["_GMV"] / df["_code预算"], np.nan)

# 折扣率 = code / 页面价
df["_折扣率"] = np.where(df["_页面价"] > 0, code_values / df["_页面价"], 0)

# 叠加补贴力度 = (百补金额 + code) / (原价 - 店铺券)
denominator = df[col_price] - coupon_values
df["_叠加补贴"] = np.where(
    denominator > 0,
    (df["_百补金额"] + code_values) / denominator,
    0,
)

# 站外美金价（如果只有韩元，自动换算）
if col_krw and col_usd_ext:
    # 优先用已有的美金价，没有则从韩元换算
    df["_站外美金"] = df[col_usd_ext].copy()
    mask_need_convert = df["_站外美金"].isna() & df[col_krw].notna()
    df.loc[mask_need_convert, "_站外美金"] = df.loc[mask_need_convert, col_krw] / exchange_rate
elif col_krw:
    df["_站外美金"] = df[col_krw] / exchange_rate
elif col_usd_ext:
    df["_站外美金"] = df[col_usd_ext]
else:
    df["_站外美金"] = np.nan

# 比价结果（站外价为0或空视为未填写，不做比价）
_ext_invalid = df["_站外美金"].isna() | (df["_站外美金"] == 0)
df["_比价结果"] = np.where(
    _ext_invalid | df["_最终价格"].isna(),
    "",
    np.where(df["_最终价格"] > df["_站外美金"], "AE价高 ⚠️",
             np.where(df["_最终价格"] < df["_站外美金"], "AE价低 ✅", "同价")),
)

# ─────────────────────────────────────────────
# Step 4: 校验
# ─────────────────────────────────────────────
errors = []
warnings = []

for idx, row in df.iterrows():
    row_num = idx + 2 + header_row  # Excel行号
    name = str(row.get(col_name, f"第{row_num}行"))[:20] if col_name else f"第{row_num}行"

    # 跳过空白行（商品名、商品ID、报名原价全为空 → 非数据行）
    _name_val = row.get(col_name) if col_name else None
    _pid_val = row.get(col_map.get("商品ID")) if col_map.get("商品ID") else None
    _price_check = row.get(col_price) if col_price else None
    if (pd.isna(_name_val) or str(_name_val).strip() == "") and \
       (pd.isna(_pid_val) or str(_pid_val).strip() == "") and \
       (pd.isna(_price_check) or _price_check == 0):
        continue

    # 校验1: code金额必须为整数或.5
    if check_code_round:
        if auto_code:
            code_val = code_values.get(idx, 0)
        elif col_code:
            code_val = row.get(col_code)
        else:
            code_val = 0
        if pd.notna(code_val) and code_val != 0:
            remainder = code_val % 0.5
            if abs(remainder) > 0.001 and abs(remainder - 0.5) > 0.001:
                errors.append(f"行{row_num} [{name}]: code金额 {code_val} 不是整数或.5")

    # 校验2: 叠加补贴不能超上限
    subsidy = row.get("_叠加补贴", 0)
    if pd.notna(subsidy) and subsidy > total_cap_rate + 0.001:
        errors.append(
            f"行{row_num} [{name}]: 叠加补贴 {subsidy:.1%} 超过上限 {total_cap_rate:.0%}"
        )

    # 校验3: 比价 - AE价高
    if check_price_vs_external:
        result = row.get("_比价结果", "")
        if "AE价高" in str(result):
            warnings.append(
                f"行{row_num} [{name}]: AE最终价 ${row['_最终价格']:.2f} > 站外 ${row['_站外美金']:.2f}"
            )

    # 校验4: 报名原价异常（非数字或为0）
    price_val = row.get(col_price)
    if pd.isna(price_val) or price_val == 0:
        errors.append(f"行{row_num} [{name}]: 报名原价为空或为0")

    # 校验5: 最终价格为负
    final_price = row.get("_最终价格")
    if pd.notna(final_price) and final_price < 0:
        errors.append(f"行{row_num} [{name}]: 最终价格为负 ({final_price:.2f})")

    # 校验6: 报名原价异常高（疑似韩币误填入美金列）
    if pd.notna(price_val) and price_val > 2000:
        corrected = price_val / exchange_rate
        errors.append(
            f"行{row_num} [{name}]: 报名原价 {price_val:,.0f} 异常高，"
            f"疑似韩币误填入美金列（若为韩币则约 ${corrected:.2f}）"
        )
    elif pd.notna(price_val) and price_val > 500:
        corrected = price_val / exchange_rate
        warnings.append(
            f"行{row_num} [{name}]: 报名原价 ${price_val:,.0f} 较高，"
            f"请确认是美金（若为韩币则约 ${corrected:.2f}）"
        )

# ── 上传页签：列识别结果 ──
with tab_upload:
    # 显示识别结果
    with st.expander("📋 列识别结果（点击展开检查）", expanded=False):
        identified = len(col_map)
        total = len(COLUMN_ALIASES)
        st.write(f"已识别 {identified}/{total} 列")
        for key, col in col_map.items():
            st.write(f"  ✅ {key} → `{col}`")
        missing = set(COLUMN_ALIASES.keys()) - set(col_map.keys())
        if missing:
            st.write(f"  ⚠️ 未识别: {', '.join(missing)}")

# ── 计算页签：概览指标 + 结果表 + 站外比价 ──
with tab_calc:
    # ─────────────────────────────────────────────
    # Step 5: 展示结果
    # ─────────────────────────────────────────────
    st.markdown("---")

    # 概览指标（排除空白行）
    col1, col2, col3, col4, col5 = st.columns(5)
    # 有效行：商品名/商品ID/报名原价 至少有一个非空
    _valid_mask = pd.Series(False, index=df.index)
    if col_name:
        _valid_mask |= df[col_name].notna() & (df[col_name].astype(str).str.strip() != "")
    if col_map.get("商品ID"):
        _valid_mask |= df[col_map["商品ID"]].notna() & (df[col_map["商品ID"]].astype(str).str.strip() != "")
    if col_price:
        _valid_mask |= df[col_price].notna() & (df[col_price] != 0)
    df_valid = df[_valid_mask]

    total_items = len(df_valid)
    total_gmv = df_valid["_GMV"].sum()
    total_budget = df_valid["_code预算"].sum()
    avg_roi = df_valid["_ROI"].mean() if df_valid["_ROI"].notna().any() else 0
    price_high_count = df_valid["_比价结果"].str.contains("AE价高", na=False).sum()

    col1.metric("商品数", f"{total_items}")
    col2.metric("预估总GMV", f"${total_gmv:,.0f}")
    col3.metric("Code总预算", f"${total_budget:,.0f}")
    col4.metric("平均ROI", f"{avg_roi:.1f}x")
    col5.metric("AE价高商品", f"{price_high_count} 个", delta=f"-{price_high_count}" if price_high_count > 0 else "0", delta_color="inverse")
    # ── 筛选栏 ──
    st.markdown("---")
    st.markdown("### 🔍 筛选结果表")

    col_channel = col_map.get("频道名")
    col_owner = col_map.get("负责人")

    # 频道名向前填充（可能只在第一行出现）
    if col_channel:
        df["_频道名_filled"] = df[col_channel].ffill()

    df_view = df.copy()

    f1, f2, f3, f4 = st.columns(4)
    with f1:
        if col_channel:
            channels = df["_频道名_filled"].dropna().unique().tolist()
            sel_channel = st.selectbox("网红频道", ["全部"] + channels, key="flt_channel")
            if sel_channel != "全部":
                df_view = df_view[df_view["_频道名_filled"] == sel_channel]
    with f2:
        if col_supply:
            supplies = df_view[col_supply].dropna().astype(str).unique().tolist()
            sel_supply = st.selectbox("供给类型", ["全部"] + supplies, key="flt_supply")
            if sel_supply != "全部":
                df_view = df_view[df_view[col_supply].astype(str) == sel_supply]
    with f3:
        if col_brand:
            brand_vals = df_view[col_brand].dropna().astype(str).unique().tolist()
            sel_brand = st.selectbox("Brand+", ["全部"] + brand_vals, key="flt_brand")
            if sel_brand != "全部":
                df_view = df_view[df_view[col_brand].astype(str) == sel_brand]
    with f4:
        if col_name:
            kw = st.text_input("搜索商品名", key="flt_kw", placeholder="输入关键词")
            if kw.strip():
                df_view = df_view[df_view[col_name].astype(str).str.contains(kw.strip(), na=False, case=False)]

    st.caption(f"筛选后 {len(df_view)} 条记录（共 {len(df)} 条）")

    # 重合商品检测
    col_pid = col_map.get("商品ID")
    if col_pid and col_channel:
        product_counts = df.groupby(col_pid)[col_pid].transform("count")
        df["_重合"] = product_counts > 1
        overlap_count = df["_重合"].sum()
        if overlap_count > 0:
            st.info(f"📌 有 {overlap_count} 条记录涉及重合商品（同一商品被多个网红选中）")

    # 展示计算结果表
    st.markdown("### 📋 计算结果")

    # 构建展示用的DataFrame
    display_cols = {}
    if col_name:
        display_cols["商品名"] = df_view[col_name]
    if col_pid:
        display_cols["商品ID"] = df_view[col_pid]
    if col_map.get("SKU"):
        display_cols["SKU"] = df_view[col_map["SKU"]]
    if col_qty:
        display_cols["数量"] = df_view[col_qty]
    if col_supply:
        display_cols["供给类型"] = df_view[col_supply]
    if col_brand:
        display_cols["Brand+"] = df_view[col_brand]

    display_cols["报名原价($)"] = df_view[col_price]
    display_cols["百补金额($)"] = df_view["_百补金额"].round(2)
    display_cols["页面价($)"] = df_view["_页面价"].round(2)
    if col_coupon:
        display_cols["店铺券($)"] = df_view[col_coupon]
    display_cols["Code金额($)"] = code_values.loc[df_view.index]
    display_cols["最终价格($)"] = df_view["_最终价格"].round(2)
    display_cols["Code预算($)"] = df_view["_code预算"].round(0)
    display_cols["GMV($)"] = df_view["_GMV"].round(0)
    display_cols["ROI"] = df_view["_ROI"].round(2)
    display_cols["叠加补贴"] = (df_view["_叠加补贴"] * 100).round(1).astype(str) + "%"
    display_cols["站外价($)"] = df_view["_站外美金"].round(2)
    display_cols["比价结果"] = df_view["_比价结果"]

    df_display = pd.DataFrame(display_cols)

    # 使用data_editor展示（可编辑）
    edited_df = st.data_editor(
        df_display,
        width="stretch",
        hide_index=True,
        num_rows="fixed",
        column_config={
            "报名原价($)": st.column_config.NumberColumn(format="%.2f"),
            "百补金额($)": st.column_config.NumberColumn(format="%.2f"),
            "页面价($)": st.column_config.NumberColumn(format="%.2f"),
            "最终价格($)": st.column_config.NumberColumn(format="%.2f"),
            "Code预算($)": st.column_config.NumberColumn(format="%.0f"),
            "GMV($)": st.column_config.NumberColumn(format="%.0f"),
            "ROI": st.column_config.NumberColumn(format="%.2f"),
            "站外价($)": st.column_config.NumberColumn(format="%.2f"),
        },
    )
    # ─────────────────────────────────────────────
    # Step 5.5: Naver 站外比价（可选）
    # ─────────────────────────────────────────────
    st.markdown("---")

    if not enable_naver:
        st.markdown("### 🔎 站外比价（已跳过）")
        st.info(
            "Naver 站外比价未启用。当前仅输出价格链路计算结果。"
            "如需完整比价报告（站外最低价 + 链接），请在左侧勾选「启用 Naver 站外比价」。"
        )
    else:
        st.markdown("### 🔎 Naver 站外比价")
        st.caption(
            "用最终优惠价对比韩国站外最低价。搜索词 = Brand + 商品名核心词 + SKU选项。"
            "若Naver最低是AE链接→标红并再搜站外；若最低是站外→直接记录。"
        )

    # Naver 比价工具函数
    NAVER_BLOCK_WORDS = ["중고", "리퍼", "박스훼손", "렌탈", "중고나라", "당근", "번개장터"]
    AE_DOMAINS = ["aliexpress.com", "aliexpress.us", "aliexpress.ru", "aliexpress.io"]


    def _clean_html(raw_html):
        return re.sub(r"<.*?>", "", raw_html)


    def _is_ae_link(url):
        url_lower = str(url).lower()
        return any(d in url_lower for d in AE_DOMAINS)


    def _build_naver_query(brand, product_name, sku_option):
        """Brand + 商品名核心词(前4) + SKU选项"""
        parts = []
        if pd.notna(brand) and str(brand).strip():
            parts.append(str(brand).strip())
        if pd.notna(product_name):
            generic = ["야외", "캠핑", "피크닉", "여행", "휴대용", "다기능",
                       "스테인리스", "스틸", "대용량", "경량", "방수", "미니"]
            words = str(product_name).strip().split()
            core = [w for w in words if w not in generic]
            parts.extend(core[:4])
        if pd.notna(sku_option):
            sku = str(sku_option).strip()
            if sku and sku not in ("单一sku", "단일sku", "nan", ""):
                parts.append(sku)
        query = " ".join(parts)
        return query[:60] if len(query) > 60 else query


    def _search_naver_shop(query, cid, csecret, exclude_ae=False):
        """调 Naver Shopping API，返回 (results_list, error_str)"""
        encoded = quote(query)
        url = f"https://openapi.naver.com/v1/search/shop.json?query={encoded}&display=20&sort=asc"
        headers = {"X-Naver-Client-Id": cid, "X-Naver-Client-Secret": csecret}
        try:
            resp = requests.get(url, headers=headers, timeout=10)
            if resp.status_code != 200:
                return [], f"API错误({resp.status_code})"
            items = resp.json().get("items", [])
            results = []
            for item in items:
                title = _clean_html(item.get("title", ""))
                link = item.get("link", "")
                price = int(item.get("lprice", 0))
                mall = item.get("mallName", "")
                if any(w in title.lower() for w in NAVER_BLOCK_WORDS):
                    continue
                if exclude_ae and _is_ae_link(link):
                    continue
                results.append({
                    "title": title,
                    "link": link,
                    "price_krw": price,
                    "price_usd": round(price / exchange_rate, 2),
                    "mall": mall,
                    "is_ae": _is_ae_link(link),
                })
            results.sort(key=lambda x: x["price_krw"])
            return results, None
        except requests.exceptions.Timeout:
            return [], "超时"
        except Exception as e:
            return [], str(e)[:40]


        # 获取 Naver 密钥
        _use_naver_id = st.session_state.get("naver_id", naver_id).strip()
        _use_naver_secret = st.session_state.get("naver_secret", naver_secret).strip()

        if not _use_naver_id or not _use_naver_secret:
            st.info("👈 请先在侧边栏填入 Naver API 密钥并保存，然后即可开始站外比价。")
        else:
            # 识别 Brand 列和 SKU 文字列
            col_brand_name = None
            for c in df.columns:
                cs = str(c).replace("\n", " ").strip()
                if "Brand" in cs or "brand" in cs:
                    col_brand_name = c
                    break

            col_sku_text = None
            for c in df.columns:
                cs = str(c).replace("\n", " ").strip()
                if ("SKU" in cs or "옵션" in cs) and "ID" not in cs:
                    col_sku_text = c
                    break

            st.caption(
                f"列识别 → Brand: `{col_brand_name or '未找到'}` | "
                f"SKU文字: `{col_sku_text or '未找到'}` | "
                f"商品名: `{col_name or '未找到'}`"
            )

            if st.button("🚀 开始 Naver 站外比价", type="primary", width="stretch"):
                # 初始化结果存储
                naver_results = []
                total = len(df)
                progress_bar = st.progress(0)
                status_text = st.empty()

                for idx, row in df.iterrows():
                    i = df.index.get_loc(idx)
                    status_text.text(f"正在比价 [{i+1}/{total}]...")
                    progress_bar.progress((i + 1) / total)

                    brand_val = row.get(col_brand_name, "") if col_brand_name else ""
                    name_val = row.get(col_name, "") if col_name else ""
                    sku_val = row.get(col_sku_text, "") if col_sku_text else ""
                    final_price = row.get("_最终价格", 0)
                    reg_price = row.get(col_price, 0)

                    query = _build_naver_query(brand_val, name_val, sku_val)

                    res_entry = {
                        "idx": idx,
                        "query": query,
                        "lowest_is_ae": False,
                        "ae_price_krw": None,
                        "ae_price_usd": None,
                        "ae_link": "",
                        "ae_title": "",
                        "reg_flag": "",
                        "ext_price_krw": None,
                        "ext_price_usd": None,
                        "ext_link": "",
                        "ext_mall": "",
                        "ext_title": "",
                        "vs_final": "",
                        "error": "",
                    }

                    # 原表已有站外数据 → 跳过抓取，保留原值
                    existing_krw = row.get(col_krw) if col_krw else None
                    existing_usd = row.get(col_usd_ext) if col_usd_ext else None
                    has_existing = (pd.notna(existing_krw) and existing_krw not in ("", 0)) or \
                                   (pd.notna(existing_usd) and existing_usd not in ("", 0))
                    if has_existing:
                        res_entry["ext_price_krw"] = existing_krw if pd.notna(existing_krw) else None
                        res_entry["ext_price_usd"] = existing_usd if pd.notna(existing_usd) else None
                        res_entry["error"] = "原表已有，跳过"
                        naver_results.append(res_entry)
                        continue

                    if not query.strip():
                        res_entry["error"] = "无搜索词"
                        naver_results.append(res_entry)
                        continue

                    # 第一轮搜索（含AE）
                    all_res, err = _search_naver_shop(query, _use_naver_id, _use_naver_secret, exclude_ae=False)
                    if err:
                        res_entry["error"] = err
                        naver_results.append(res_entry)
                        time.sleep(0.3)
                        continue

                    if not all_res:
                        res_entry["error"] = "无匹配"
                        naver_results.append(res_entry)
                        time.sleep(0.3)
                        continue

                    lowest = all_res[0]

                    if lowest["is_ae"]:
                        # 最低是AE → 记录 + 判断报名价 + 再搜站外
                        res_entry["lowest_is_ae"] = True
                        res_entry["ae_price_krw"] = lowest["price_krw"]
                        res_entry["ae_price_usd"] = lowest["price_usd"]
                        res_entry["ae_link"] = lowest["link"]
                        res_entry["ae_title"] = lowest["title"][:40]

                        if pd.notna(reg_price) and reg_price > 0 and lowest["price_usd"] < reg_price:
                            res_entry["reg_flag"] = "⚠️ AE实际价<报名价"

                        # 第二轮：排除AE搜站外
                        time.sleep(0.3)
                        ext_res, ext_err = _search_naver_shop(query, _use_naver_id, _use_naver_secret, exclude_ae=True)
                        if ext_res:
                            ext_low = ext_res[0]
                            res_entry["ext_price_krw"] = ext_low["price_krw"]
                            res_entry["ext_price_usd"] = ext_low["price_usd"]
                            res_entry["ext_link"] = ext_low["link"]
                            res_entry["ext_mall"] = ext_low["mall"]
                            res_entry["ext_title"] = ext_low["title"][:40]
                    else:
                        # 最低是站外 → 直接记录
                        res_entry["ext_price_krw"] = lowest["price_krw"]
                        res_entry["ext_price_usd"] = lowest["price_usd"]
                        res_entry["ext_link"] = lowest["link"]
                        res_entry["ext_mall"] = lowest["mall"]
                        res_entry["ext_title"] = lowest["title"][:40]

                    # 最终价 vs 站外最低
                    if res_entry["ext_price_usd"] and pd.notna(final_price) and final_price > 0:
                        if final_price <= res_entry["ext_price_usd"]:
                            res_entry["vs_final"] = "✅ AE价低"
                        else:
                            diff_pct = (final_price - res_entry["ext_price_usd"]) / res_entry["ext_price_usd"] * 100
                            if diff_pct <= tolerance_pass_pct:
                                res_entry["vs_final"] = f"≈ 持平(+{diff_pct:.0f}%)"
                            elif diff_pct <= tolerance_warn_pct:
                                res_entry["vs_final"] = f"⚠️ AE略高(+{diff_pct:.0f}%)"
                            else:
                                res_entry["vs_final"] = f"❌ AE价高(+{diff_pct:.0f}%)"

                    naver_results.append(res_entry)
                    time.sleep(0.3)

                # 存入 session_state 以便 rerun 后仍可查看
                st.session_state["naver_results"] = naver_results
                status_text.text(f"比价完成！共 {total} 个商品")
                progress_bar.progress(1.0)

            # 展示 Naver 比价结果
            if "naver_results" in st.session_state:
                naver_results = st.session_state["naver_results"]

                # 统计
                n_ok = sum(1 for r in naver_results if "AE价低" in r.get("vs_final", ""))
                n_warn = sum(1 for r in naver_results if "略高" in r.get("vs_final", "") or "持平" in r.get("vs_final", ""))
                n_bad = sum(1 for r in naver_results if "AE价高" in r.get("vs_final", ""))
                n_ae_lowest = sum(1 for r in naver_results if r.get("lowest_is_ae"))
                n_flag = sum(1 for r in naver_results if r.get("reg_flag"))
                n_err = sum(1 for r in naver_results if r.get("error"))

                nc1, nc2, nc3, nc4, nc5 = st.columns(5)
                nc1.metric("✅ AE价低", n_ok)
                nc2.metric("⚠️ 略高/持平", n_warn)
                nc3.metric("❌ AE价高", n_bad)
                nc4.metric("AE是Naver最低", n_ae_lowest)
                nc5.metric("查询失败", n_err)

                if n_flag > 0:
                    st.warning(f"⚠️ {n_flag} 个商品：AE实际价 < 行业报名价（报名价可能虚高）")

                # 构建结果表
                naver_display = []
                for r in naver_results:
                    row_data = {
                        "搜索词": r["query"][:30],
                        "最低是AE": "是" if r["lowest_is_ae"] else "否",
                        "AE价(₩)": r["ae_price_krw"] or "",
                        "AE价($)": r["ae_price_usd"] or "",
                        "报名价标记": r["reg_flag"],
                        "站外最低(₩)": r["ext_price_krw"] or "",
                        "站外最低($)": r["ext_price_usd"] or "",
                        "站外商城": r["ext_mall"],
                        "站外链接": r["ext_link"],
                        "vs最终价": r["vs_final"],
                        "错误": r["error"],
                    }
                    naver_display.append(row_data)

                df_naver = pd.DataFrame(naver_display)

                # 筛选视图
                naver_filter = st.radio(
                    "显示",
                    ["全部", "仅看问题行（AE价高/报名价虚高）", "仅看失败"],
                    horizontal=True,
                    key="naver_filter",
                )
                if naver_filter == "仅看问题行（AE价高/报名价虚高）":
                    mask = df_naver["vs_final"].str.contains("AE价高", na=False) | (df_naver["报名价标记"] != "")
                    df_naver_show = df_naver[mask]
                elif naver_filter == "仅看失败":
                    df_naver_show = df_naver[df_naver["错误"] != ""]
                else:
                    df_naver_show = df_naver

                st.dataframe(
                    df_naver_show,
                    width="stretch",
                    hide_index=True,
                    column_config={
                        "站外链接": st.column_config.LinkColumn("打开"),
                    },
                )

                # 将 Naver 结果写回 df 以便导出
                for r in naver_results:
                    idx = r["idx"]
                    if r["ext_price_krw"]:
                        df.loc[idx, "_站外韩元"] = r["ext_price_krw"]
                        df.loc[idx, "_站外美金"] = r["ext_price_usd"]
                    if r["ext_link"]:
                        df.loc[idx, "_站外链接"] = r["ext_link"]
                    if r["vs_final"]:
                        df.loc[idx, "_比价结果"] = r["vs_final"]

# ── 校验页签 ──
with tab_valid:
    # ── 校验仪表盘 ──
    st.markdown("### 校验概览")

    def _parse_issue(msg, level):
        m = re.match(r"行(\d+)\s*\[(.+?)\]:\s*(.+)", msg)
        if m:
            return {"级别": level, "行号": m.group(1), "商品名": m.group(2), "问题": m.group(3)}
        return {"级别": level, "行号": "", "商品名": "", "问题": msg}

    issue_rows = [_parse_issue(e, "❌ 错误") for e in errors] + \
                 [_parse_issue(w, "⚠️ 警告") for w in warnings]
    issue_row_nums = {r["行号"] for r in issue_rows if r["行号"] != ""}

    c1, c2, c3 = st.columns(3)
    c1.metric("❌ 错误", f"{len(errors)}")
    c2.metric("⚠️ 警告", f"{len(warnings)}")
    c3.metric("📋 涉及商品", f"{len(issue_row_nums)} 个")

    st.markdown("### 问题明细")
    if issue_rows:
        df_issues_all = pd.DataFrame(issue_rows)
        lvl_filter = st.radio("显示", ["全部", "仅错误", "仅警告"], horizontal=True, key="valid_filter")
        if lvl_filter == "仅错误":
            df_issue_show = df_issues_all[df_issues_all["级别"].str.contains("错误")]
        elif lvl_filter == "仅警告":
            df_issue_show = df_issues_all[df_issues_all["级别"].str.contains("警告")]
        else:
            df_issue_show = df_issues_all
        st.dataframe(df_issue_show, width="stretch", hide_index=True, height=380)
    else:
        st.success("✅ 全部校验通过，无异常！")

# ── 分析页签：GMV 分布 ──
with tab_analyze:
    # ── GMV 分布预览 ──
    # 处理品牌文件
    _brand_df = None
    if brand_file:
        try:
            _brand_df = pd.read_excel(brand_file)
            # 标准化列名
            _brand_cols = {}
            for c in _brand_df.columns:
                cs = str(c).replace("\n", " ").strip()
                if "Brand" in cs:
                    _brand_cols["brand"] = c
                elif "商品ID" in cs or "상품ID" in cs:
                    _brand_cols["pid"] = c
                elif "一级类目" in cs or "1급" in cs:
                    _brand_cols["cat1"] = c
            if "pid" in _brand_cols and col_map.get("商品ID"):
                _brand_df["_pid_str"] = _brand_df[_brand_cols["pid"]].astype(str).str.strip()
                _brand_df = _brand_df.rename(columns={
                    _brand_cols.get("brand", "_none"): "_brand",
                    _brand_cols.get("cat1", "_none"): "_cat1",
                })
        except Exception:
            _brand_df = None

    # 构建分析用df
    _gmv_df = df_valid.copy()
    _gmv_df["_gmv"] = _gmv_df["_GMV"]

    # 如果有品牌文件，join品牌信息
    _has_brand = False
    if _brand_df is not None and "pid" in _brand_cols and col_map.get("商品ID"):
        _pid_col = col_map["商品ID"]
        _gmv_df["_pid_str"] = _gmv_df[_pid_col].astype(str).str.strip()
        _merge_cols = ["_pid_str"]
        if "_brand" in _brand_df.columns:
            _merge_cols.append("_brand")
        if "_cat1" in _brand_df.columns:
            _merge_cols.append("_cat1")
        _gmv_df = _gmv_df.merge(
            _brand_df[_merge_cols].drop_duplicates(subset="_pid_str"),
            on="_pid_str", how="left",
        )
        _gmv_df["_brand"] = _gmv_df.get("_brand", pd.Series()).fillna("未知品牌")
        _gmv_df["_brand"] = _gmv_df["_brand"].replace({"#REF!": "未知品牌", "": "未知品牌"})
        _has_brand = True

    st.markdown("### GMV 分布")

    # 维度选择
    dim_options = ["按网红"]
    if _has_brand:
        dim_options = ["按品牌", "按网红", "按一级类目"]
    dim = st.radio("分析维度", dim_options, horizontal=True, key="gmv_dim")

    # 确定分组列
    if dim == "按品牌" and _has_brand:
        _group_col = "_brand"
        _group_label = "品牌"
    elif dim == "按一级类目" and _has_brand and "_cat1" in _gmv_df.columns:
        _group_col = "_cat1"
        _group_label = "一级类目"
    else:
        _group_col = col_map.get("频道名")
        _group_label = "网红"
        if _group_col:
            _gmv_df[_group_col] = _gmv_df[_group_col].ffill()

    if _group_col and _group_col in _gmv_df.columns:
        # 汇总
        _summary = _gmv_df.groupby(_group_col).agg(
            SKU数=("_gmv", "count"),
            预估GMV=("_gmv", "sum"),
        ).reset_index()
        _summary = _summary.rename(columns={_group_col: _group_label})
        _summary = _summary.sort_values("预估GMV", ascending=False).reset_index(drop=True)
        _total_gmv = _summary["预估GMV"].sum()
        _summary["占比"] = (_summary["预估GMV"] / _total_gmv * 100).round(1).astype(str) + "%"

        # 如果有code预算列，加上
        if "_code预算" in _gmv_df.columns:
            _budget = _gmv_df.groupby(_group_col)["_code预算"].sum()
            _summary["Code预算"] = _summary[_group_label].map(_budget).fillna(0)

        # ── 集中度分析 ──
        _gmv_sorted = _summary["预估GMV"].sort_values(ascending=False)
        _top3_pct = _gmv_sorted.head(3).sum() / _total_gmv * 100 if _total_gmv > 0 else 0
        _top5_pct = _gmv_sorted.head(5).sum() / _total_gmv * 100 if _total_gmv > 0 else 0
        cm1, cm2, cm3 = st.columns(3)
        cm1.metric(f"{_group_label}总数", f"{len(_summary)}")
        cm2.metric("Top3 集中度", f"{_top3_pct:.1f}%")
        cm3.metric("Top5 集中度", f"{_top5_pct:.1f}%")

        # ── 左控制 · 右可视化 ──
        ctrl_col, viz_col = st.columns([1, 3])
        with ctrl_col:
            view_mode = st.radio("展示形式", ["柱状图", "表格"], key="gmv_view")
            if len(_summary) > 5:
                top_n = st.slider("显示 Top N", min_value=5, max_value=min(50, len(_summary)), value=min(15, len(_summary)), key="gmv_topn")
            else:
                top_n = len(_summary)
        _show = _summary.head(top_n)
        with viz_col:
            if view_mode == "表格":
                st.dataframe(_show, width="stretch", hide_index=True)
            else:
                _chart_data = _show.set_index(_group_label)["预估GMV"].sort_values(ascending=True)
                st.bar_chart(_chart_data, horizontal=True, color="#4472C4")

        st.caption(f"共 {len(_summary)} 个{_group_label}，总预估GMV ${_total_gmv:,.0f}")
    else:
        st.info("未识别到网红频道名列，无法生成分布。")

# ── 导出页签 ──
with tab_export:
    # ─────────────────────────────────────────────
    # Step 6: 导出（openpyxl 格式化）
    # ─────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📥 导出")




    col_a, col_b = st.columns(2)

    with col_a:
        # 构建导出DataFrame
        export_df = df.copy()

        # 把计算列写回原表列名
        if col_map.get("百补金额"):
            export_df[col_map["百补金额"]] = export_df["_百补金额"]
        if col_map.get("百补力度"):
            # R列：直接用已计算的_百补力度（全托管→5%，半托POP非brand+→0不填）
            _rate_vals = export_df["_百补力度"] if "_百补力度" in export_df.columns else None
            if _rate_vals is not None:
                export_df.loc[_rate_vals > 0, col_map["百补力度"]] = _rate_vals[_rate_vals > 0]
        if col_map.get("页面价"):
            export_df[col_map["页面价"]] = export_df["_页面价"]
        if col_map.get("最终价格"):
            export_df[col_map["最终价格"]] = export_df["_最终价格"]
        if col_map.get("code金额"):
            # Y列：覆盖为工具计算值（用户确认方案A）
            export_df[col_map["code金额"]] = code_values.values
        if col_map.get("code预算"):
            export_df[col_map["code预算"]] = export_df["_code预算"]
        if col_map.get("GMV"):
            export_df[col_map["GMV"]] = export_df["_GMV"]
        if col_map.get("ROI"):
            export_df[col_map["ROI"]] = export_df["_ROI"]
        if col_map.get("叠加补贴力度"):
            export_df[col_map["叠加补贴力度"]] = export_df["_叠加补贴"]
        if col_map.get("折扣率"):
            export_df[col_map["折扣率"]] = export_df["_折扣率"]
        if col_result:
            export_df[col_result] = export_df["_比价结果"]

        # Naver 比价结果写回
        if col_map.get("站外美金") and "_站外美金" in export_df.columns:
            export_df[col_map["站外美金"]] = export_df["_站外美金"]
        if col_map.get("站外韩元") and "_站外韩元" in export_df.columns:
            export_df[col_map["站外韩元"]] = export_df["_站外韩元"]
        if col_map.get("站外链接") and "_站外链接" in export_df.columns:
            export_df[col_map["站外链接"]] = export_df["_站外链接"]
        elif "_站外链接" in export_df.columns:
            export_df["站外比价链接"] = export_df["_站外链接"]

        # 去掉内部辅助列
        internal_cols = [c for c in export_df.columns if c.startswith("_")]
        export_df = export_df.drop(columns=internal_cols)

        # 生成格式化Excel
        buffer = build_formatted_excel(export_df, col_map.get("频道名"))

        st.download_button(
            label="📥 下载完整价格链路表（含计算结果）",
            data=buffer,
            file_name="价格链路_计算完成.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )

    with col_b:
        # 导出异常商品清单
        if errors or warnings:
            issues_data = []
            for e in errors:
                issues_data.append({"类型": "❌ 错误", "详情": e})
            for w in warnings:
                issues_data.append({"类型": "⚠️ 警告", "详情": w})
            df_issues = pd.DataFrame(issues_data)

            buffer2 = BytesIO()
            df_issues.to_excel(buffer2, index=False, engine="openpyxl")
            buffer2.seek(0)

            st.download_button(
                label="📥 下载异常清单",
                data=buffer2,
                file_name="价格链路_异常清单.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )
    # ─────────────────────────────────────────────
    # 过往价格对比（从商品名中提取）
    # ─────────────────────────────────────────────
    if col_name:
        st.markdown("---")
        st.markdown("### 📈 过往价格对比")
        st.caption("自动从商品名中提取「X月最终价」记录，与本次最终价对比")

        history_records = []
        for idx, row in df.iterrows():
            name = str(row.get(col_name, ""))
            # 匹配 "3月最终价：30.98" 或 "3月最终价30.98" 或 "6月最终价：73.33"
            match = re.search(r"(\d+)月最终价[：:]?\s*([\d.]+)", name)
            if match:
                month = match.group(1)
                old_price = float(match.group(2))
                new_price = row.get("_最终价格", np.nan)
                if pd.notna(new_price):
                    diff = new_price - old_price
                    diff_pct = (diff / old_price * 100) if old_price > 0 else 0
                    history_records.append({
                        "商品名": name[:30],
                        f"{month}月最终价": old_price,
                        "本次最终价": round(new_price, 2),
                        "差额": round(diff, 2),
                        "涨幅": f"{diff_pct:+.1f}%",
                        "状态": "⚠️ 涨价" if diff > 0.5 else ("✅ 降价" if diff < -0.5 else "≈ 持平"),
                    })

        if history_records:
            df_history = pd.DataFrame(history_records)
            st.dataframe(df_history, width="stretch", hide_index=True)
            price_up = sum(1 for r in history_records if "涨价" in r["状态"])
            if price_up > 0:
                st.warning(f"有 {price_up} 个商品本次价格高于过往，建议关注")
        else:
            st.info("未在商品名中发现过往价格记录。")
# 页脚
st.markdown("---")
st.caption("网红团购一站式平台 · 价格链路模块 v1.2 | 数据仅在本地浏览器处理，不上传任何服务器")
